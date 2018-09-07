import pandas as pd
import numpy as np
import argparse
from openpyxl import load_workbook

parser = argparse.ArgumentParser()
parser.add_argument("courselist", help="list of courses in tsv format")
parser.add_argument("iuaexlsx", help="xlsx document that containts the IUAE matrix")
parser.add_argument("coursecode", help="course code")
args = parser.parse_args()

tiuae = pd.read_csv(args.courselist, sep='\t')
wbiuae = load_workbook(args.iuaexlsx)
ws = wbiuae.active

res = tiuae[tiuae.Kod == args.coursecode]
nn = res.iat[0, 1]
nc = res.iat[0, 2]

ws['A6'] = f'{nn}, {nc} hp'

wbiuae.save(args.iuaexlsx)