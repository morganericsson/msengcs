import pandas as pd
import numpy as np
import argparse
from openpyxl import load_workbook

parser = argparse.ArgumentParser()
parser.add_argument("iuaetemplate", help="the template to use for the output tsv")
parser.add_argument("iuaexlsx", help="xlsx document that containts the IUAE matrix")
parser.add_argument("iuaetsv", help="tsv file to store the IUAE matrix in")
args = parser.parse_args()

tiuae = pd.read_csv(args.iuaetemplate, sep='\t', dtype={'CDIO':str})
wbiuae = load_workbook(args.iuaexlsx)
ws = wbiuae.active

fields = { 'Introducera':'C', 'Undervisa':'D', 'Använda':'E', 'Examinera':'F', 'Kommentar':'G' }

for f,c in fields.items():
	for i in range(8,35):
		tiuae[f].loc[i-7] = ws[f'{c}{i}'].value

tiuae.to_csv(args.iuaetsv, sep='\t', index=False)